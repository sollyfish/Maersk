import streamlit as st
import geopandas as gpd
import matplotlib.pyplot as plt
import pandas as pd
import matplotlib.patches as mpatches
import os
from openpyxl import load_workbook
from io import BytesIO
from copy import copy

# ---------------- Resource path (repo-root safe) ----------------
def resource_path(relative_path: str) -> str:
    return os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        "..",
        relative_path
    )

# ---------------- Cached ZIP3 shapes (GPKG) ----------------
@st.cache_resource
def load_zip3_shapes():
    gdf = gpd.read_file(
        resource_path("shapefiles/zip3_simplified.gpkg"),
        engine="pyogrio"  # more stable on Streamlit Cloud
    )
    gdf["zip3"] = gdf["zip3"].astype(str).str.zfill(3)
    return gdf

# ---------------- Heavy processing function ----------------
def process_data(origin_list, customer_name):
    progress_text = st.empty()

    # Step 1: Load Excel
    progress_text.info("Loading Excel file...")
    excel_path = resource_path("assets/Maersk Zones.xlsx")
    MasterZone_df = pd.read_excel(excel_path)

    # Step 2: Process Data
    progress_text.info("Processing zone data...")
    MasterZone_df["OriginZip"] = MasterZone_df["Set_ID"].astype(str).str.zfill(3)
    MasterZone_df["DestZipMin"] = MasterZone_df["Min_Zip_Int"].astype(int)
    MasterZone_df["DestZipMax"] = MasterZone_df["Max_Zip_Int"].astype(int)
    MasterZone_df["Zone"] = MasterZone_df["Zone"].astype(int)

    filtered = MasterZone_df[
        MasterZone_df["OriginZip"].isin(origin_list)
    ].copy()

    filtered["DestZipRange"] = [
        list(range(start, end + 1))
        for start, end in zip(filtered["DestZipMin"], filtered["DestZipMax"])
    ]

    expanded_df = filtered.explode("DestZipRange")
    expanded_df["zip3"] = expanded_df["DestZipRange"].astype(str).str.zfill(3)
    expanded_df = expanded_df[["zip3", "Zone", "OriginZip"]]

    min_zones = expanded_df.groupby("zip3")["Zone"].min().reset_index()
    min_zone_df = expanded_df.merge(min_zones, on=["zip3", "Zone"])

    expanded_df = min_zone_df.copy()

    # Step 3: Load ZIP3 shapes
    progress_text.info("Loading ZIP3 map shapes...")
    zip3_shapes = load_zip3_shapes()

    zip3_shapes = zip3_shapes.merge(expanded_df, on="zip3", how="left")

    # Step 4: Plot
    progress_text.info("Rendering map...")
    zone_colors = {
        1: "#001624", 2: "#00243D", 3: "#004A73",
        4: "#0073AB", 5: "#2392BE", 6: "#42B0D5",
        7: "#72C8E3", 8: "#A1D8EF", 9: "#B5E0F5"
    }

    fig, ax = plt.subplots(figsize=(15, 10))

    # State boundaries (lightweight, OK to reload)
    states = gpd.read_file(
        resource_path("shapefiles/states_preprocessed.gpkg"),
        engine="pyogrio"
    )
    states.boundary.plot(ax=ax, linewidth=0.5, edgecolor="black")

    ax.set_facecolor("#a3a3a3")

    zip3_plot_colors = zip3_shapes["Zone"].map(zone_colors).fillna("#CCCCCC")
    zip3_shapes.plot(ax=ax, color=zip3_plot_colors, linewidth=0)

    # Continental US view
    ax.set_xlim(-130, -65)
    ax.set_ylim(24, 50)
    ax.set_aspect(1.2, adjustable="box")

    used_zones = sorted(
        z for z in zip3_shapes["Zone"].dropna().unique()
        if z != 9
    )

    legend_handles = [
        mpatches.Patch(color=zone_colors[z], label=str(z))
        for z in used_zones
    ]
    
    ax.legend(
        handles=legend_handles,
        title="Zone",
        loc="lower left",
        fontsize="small"
    )

    ax.set_title(f"Zone Map – {customer_name}", fontsize=16)
    ax.axis("off")
    plt.tight_layout()

    progress_text.success("Done!  Now Loading Customer Zone List")

    return fig, expanded_df


def process_export_data(origin_list_3):

    excel_path = resource_path("assets/Maersk Zones.xlsx")
    df = pd.read_excel(excel_path)

    df["OriginZip"] = df["Set_ID"].astype(str).str.zfill(3)
    df["DestZip3Min"] = df["Min_Zip_Int"].astype(int)
    df["DestZip3Max"] = df["Max_Zip_Int"].astype(int)
    df["Zone"] = df["Zone"].astype(int)

    # Filter to selected origins
    df = df[df["OriginZip"].isin(origin_list_3)].copy()

    # Expand ZIP3 ranges ONLY
    df["zip3_range"] = df.apply(
        lambda r: range(r["DestZip3Min"], r["DestZip3Max"] + 1),
        axis=1
    )

    df = df.explode("zip3_range")

    df["zip3"] = df["zip3_range"].astype(str).str.zfill(3)

    return df[["zip3", "OriginZip", "Zone"]]
# ---------------- Streamlit Feature Entry Point ----------------
def zone_map_app():
    st.header("Zone Map & List Generator")

    origin_input = st.text_input(
        "Enter 5-Digit Origin ZIPs (comma separated)"
    )
    customer_name = st.text_input("Customer Name")

    if st.button("Generate Map"):
        # Step 1: Clean raw input
        raw_origins = [
            o.strip()
            for o in origin_input.split(",")
            if o.strip().isdigit()
        ]
        
        invalid_inputs = []
        valid_origins = []

        for o in raw_origins:
            if o.isdigit() and len(o) == 5:
                valid_origins.append(o)
            else:
                invalid_inputs.append(o)

        # 🚫 If ANY invalid inputs → stop everything
        if invalid_inputs:
            st.error(
                f"All ZIP codes must be 5 digits. Invalid entries: {', '.join(invalid_inputs)} \n\n" 
                f"HINT: if you don't know the full 5 digit ZIP, you can enter the first 3 digits followed by '00' (e.g. '12300' for ZIPs starting with 123)."
            )
            return

        # ✅ Safe to proceed
        origin_list_5 = [o.zfill(5) for o in valid_origins]
        origin_list_3 = [o[:3] for o in origin_list_5]

        if not origin_list_5:
            st.error("Please enter at least one valid 5-digit Origin ZIP.")
            return

        if not customer_name:
            customer_name = "Maersk"

        with st.spinner("Processing… this may take a moment"):
            fig, expanded_df = process_data(origin_list_3, customer_name)

        st.pyplot(fig)

       # ================================
        # 📥 EXPORT USING TEMPLATE (FIXED)
        # ================================
        
        export_df = process_export_data(origin_list_3)
        
        template_path = resource_path("assets/ZoningTemplate.xlsx")
        wb = load_workbook(template_path)
        ws = wb.active
        
        # ================================
        # 🧠 BUILD MATRIX (ZIP5 x ORIGIN)
        # ================================
        
        pivot_df = export_df.groupby(["zip3", "OriginZip"])["Zone"].min().unstack()
        
        # Ensure all origins exist
        for origin in origin_list_3:
            if origin not in pivot_df.columns:
                pivot_df[origin] = None
        
        # ================================
        # TEMPLATE STRUCTURE
        # ================================
        
        header_row = 4
        data_start_row = 5
        origin_start_col = 2  # Column B
        
        # ================================
        # 🎨 COPY TEMPLATE FORMATTING
        # ================================
        
        def copy_column_format(ws, source_col, target_col):
            max_row = ws.max_row

            for row in range(header_row, max_row + 1):
                source_cell = ws.cell(row=row, column=source_col)
                target_cell = ws.cell(row=row, column=target_col)
        
                if source_cell.has_style:
                    target_cell._style = copy(source_cell._style)
        
        existing_origin_cols = ws.max_column - (origin_start_col - 1)
        
        if len(origin_list_5) > existing_origin_cols:
            for i in range(existing_origin_cols, len(origin_list_5)):
                source_col = origin_start_col + existing_origin_cols - 1
                target_col = origin_start_col + i
                copy_column_format(ws, source_col, target_col)
        
        # ================================
        # 🧹 CLEAR OLD DATA (ONLY B+)
        # ================================
        
        for row in ws.iter_rows(min_row=data_start_row, max_row=ws.max_row, min_col=origin_start_col):
            for cell in row:
                cell.value = None
        
        # ================================
        # 🧹 CLEAR OLD HEADERS
        # ================================
        
        for col in range(origin_start_col, ws.max_column + 1):
            ws.cell(row=header_row, column=col).value = None
        
        # ================================
        # ✍️ WRITE HEADERS (ROW 4)
        # ================================
        
        for col_idx, origin in enumerate(origin_list_5, start=origin_start_col):
            ws.cell(row=header_row, column=col_idx, value=origin)
        
        # ================================
        # 📍 MATCH TEMPLATE ZIP5 ROWS
        # ================================
        
        template_rows = []
        
        for r in range(data_start_row, ws.max_row + 1):
            val = ws.cell(row=r, column=1).value
        
            if val is None or str(val).strip() == "":
                break  # stop at end of template
        
            zip5 = str(val).strip().zfill(5)
            template_rows.append((r, zip5))
                
        # ================================
        # ✍️ WRITE DATA (DO NOT TOUCH COL A)
        # ================================
        columns_with_data = set()

        for r_idx, zip5 in template_rows:
        
            zip3 = zip5[:3]
        
            if zip3 in pivot_df.index:
                row_data = pivot_df.loc[zip3]
        
                for c_idx, origin5 in enumerate(origin_list_5, start=origin_start_col):
        
                    origin3 = origin5[:3]
                    value = row_data.get(origin3)
        
                    if pd.notna(value):
                        ws.cell(row=r_idx, column=c_idx, value=value)
                        columns_with_data.add(c_idx)
        source_col = origin_start_col  # Column B

        # Loop through all origin columns that were written
        for col_idx in range(origin_start_col, origin_start_col + len(origin_list_5)):
            if col_idx == source_col:
                continue  # skip template column (B)
        
            copy_column_format(ws, source_col, col_idx)
                        

        
        # ================================
        # 💾 SAVE FOR DOWNLOAD
        # ================================
        ws.freeze_panes = "A5"
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        st.download_button(
            label="📥 Download Zoning Table (Template)",
            data=output,
            file_name=f"{customer_name}_zoning_table.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
